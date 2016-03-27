#!usr/bin/python3.5
from openpyxl import Workbook
import random
import math

'''
    Assumptions:
    
        The chance of giving chromosome is 50% each.
'''

class Individual:

    def __init__( self, genotype ):

        if genotype[0] > genotype[1]:
            genotype = genotype[::-1]
        self.__genotype = genotype
        self.__age      = 0

    def grow_up( self ):

        self.__age += 1
        if self.__age >= life:
            return False
        return True

    def breed( self ):

        if ( self.__age - start ) % rest == 0 \
           and self.__age <= end:
            return random.choice( list( self.__genotype ) )
        return False

    def __str__( self ):

        return self.__genotype

def main():

    global start, end, life, rest
    # How many children each parent will have?
    children   = 3
    # What is the life expectancy?
    life       = 10
    # How old will they start breed (first breedable age)?
    start      = 5
    # How old will they end breed (last breedable age)?
    end        = 7
    # How often they breed (year between two breed)?
    rest       = 2
    # What are the genotypic frequencies?
    # For 2 alleles, format is [ AA, AB, BB ]
    # For more alleles, format will be:
    # [ AA, AB, AC, ..., BB, BC, ..., CC, ..., ... ]
    frequency  = [ 0.6, 0.2, 0.2 ]
    # What is the initial number of individuals?
    population = 5000
    # How many years will we simulate?
    year       = 10

    individuals = list()
    temperary   = list()
    genotype    = list()
    position    = 0
    allele      = math.sqrt( 2 * len( frequency ) + 0.25 ) - 0.5
    alleles     = [0] * int( allele )
    excel       = Workbook()
    sheet       = excel.active
    sheet.title = "Bird Species Simulator"

    for i in range( int( allele ) ):
        for j in range( i, int( allele ) ):
            genotype.append( chr( 65 + i ) + chr( 65 + j ) )

    for i in range( population ):
        individuals.append( Individual( genotype[position] ) )
        if i == int( population * \
                     sum( frequency[:position + 1] ) - 1 ):
            position += 1

    sheet['A1'], sheet['A2'] = "Children"  , children
    sheet['A3'], sheet['A4'] = "Life"      , life
    sheet['A5'], sheet['A6'] = "Start"     , start
    sheet['A7'], sheet['A8'] = "End"       , end
    sheet['B1'], sheet['B2'] = "Year"      , 0
    sheet['C1'], sheet['C2'] = "Population", population

    for i in range( len( frequency ) ):
        sheet.cell( row = 1, column = 4 + i ).value = genotype[i]
        sheet.cell( row = 2, column = 4 + i ).value = frequency[i]

    for i in range( len( genotype ) ):
        alleles[ ord( genotype[i][0] ) - 65 ]  += 0.5 * frequency[i]
        alleles[ ord( genotype[i][1] ) - 65 ]  += 0.5 * frequency[i]

    for i in range( int( allele ) ):
        sheet.cell( row = 1, \
                    column = 4 + len( frequency ) + i ).value = \
                    chr( 65 + i )
        sheet.cell( row = 2, \
                    column = 4 + len( frequency ) + i ).value = \
                    alleles[i]

    for i in range( 1, year + 1 ):

        print( "Working on: %i/%i, " %( i, year ) )
        
        try:
            while True:
                while True:
                    print( "( %i/%i )" %( population - \
                                          len(individuals), \
                                          population ), end = "\r" )
                    male = individuals.pop( \
                        random.randrange( len( individuals ) ) )
                    if male.grow_up():
                        temperary.append( male )
                        gene = male.breed()
                        if gene:
                            break

                while True:
                    print( "( %i/%i )" %( population - \
                                          len(individuals), \
                                          population ), end = "\r" )
                    female = individuals.pop( \
                        random.randrange( len( individuals ) ) )
                    if female.grow_up():
                        temperary.append( female )
                        gene = female.breed()
                        if gene:
                            break

                for j in range( children ):
                    temperary.append( \
                        Individual( male.breed() + female.breed() ) )

        except ValueError:
            pass

        individuals, temperary = temperary, list()
        frequency              = [0] * len( frequency )
        alleles                = [0] * int( allele )
        population             = len( individuals )

        for j in individuals:
            current = str( j )
            frequency[genotype.index(current)] += 1
            alleles[ ord( current[0] ) - 65 ]  += 0.5
            alleles[ ord( current[1] ) - 65 ]  += 0.5

        sheet.cell( row = 2 + i, column = 2 ).value = i
        sheet.cell( row = 2 + i, column = 3 ).value = population

        for j in range( len( frequency ) ):
            sheet.cell( row = 2 + i, column = 4 + j ).value = \
                        frequency[j] / population

        for j in range( len( alleles ) ):
            sheet.cell( row = 2 + i, \
                        column = 4 + len( frequency ) + j ).value = \
                        alleles[j] / population

    excel.save( "math372-project.xlsx" )

if __name__ == "__main__":
    main()
